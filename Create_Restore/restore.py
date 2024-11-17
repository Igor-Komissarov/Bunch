import json
import sys
from deep_translator import GoogleTranslator
import pandas as pd
from gen_api import *
import os
import shutil
import locale
import re
from openpyxl import load_workbook
from good_gen_json import *
import yaml

def process_text_field(field, replace_dict=None, remove_country_code=False, max_length=None):
    """Обрабатывает текстовое поле, выполняя замены и, при необходимости, удаляет коды стран."""
    if not field:
        return ''
    
    # Удаление кода страны в начале строки, если remove_country_code=True
    if remove_country_code:
        field = re.sub(r'^\([A-Z]{2,3}\d+(?:/\d+)?\)\s*', '', field)
    
    # Выполнение замен, если они указаны
    if replace_dict:
        for old, new in replace_dict.items():
            field = field.replace(old, new)

    # Ограничение длины текста, если указано max_length
    if max_length and len(field) > max_length:
        field = field[:max_length]
    
    return field.replace('<br/>', '\n')

def process_PTCC(PTCC):
    """Обрабатывает поле FNUM для извлечения PASI."""
    true_data = []
    if PTCC:
        PTCC = PTCC.replace('<br/>', '\n')
        fake_PTCC = PTCC.split()
        for i in fake_PTCC:
            i = i.replace('<br/><br/>', '').split('=')
            if i[0] == 'CC':
                name = i[1]
                true_data.append(name)
    else:
        true_data.append(None)
    return true_data

def process_CLM(field, delimiter):
    if field:
        return [field.split(delimiter)]
    return field
def check(a):
    num = 1
    if a == '':
        num = 0
    return num

def extract_patent_restore(name):
    """Извлекает информацию о патентах, обрабатывая все ключевые поля."""
    return {
        'FAN': name.get('FAN', ''),
        'EAPD': name.get('EAPD', ''),
        'PA': process_text_field(name.get('PA', ''), {'<br/>': '; '}, max_length=5000),
        'EPRD': name.get('EPRD', ''),
        'PTCC': process_PTCC(name.get('PTCC', '')),
        'CLMS': process_CLM(name.get('CLMS', ''), '</p><p>'),
        'ICLM': process_CLM(name.get('ICLM', ''), '</p><p>'),
        'IC': process_text_field(name.get('IC', ''), {'<br/>': ',\n '}, max_length=5000),
        'CTGN': process_citing_cited(name.get('CTGN', ''), '<br/>'),
        'CTN': process_citing_cited(name.get('CTN', ''), '<br/>'),
        'NPR': name.get('NPR', ''),
        'NPN': name.get('NPN', ''),
        'STDN': check(name.get('STDN', '')),
        'LIC': check(name.get('LIC', '')),
        'OPPI': check(name.get('OPPI', '')),
        'V_APL': process_v_apl(name.get('V_APL', '')),
        'PASI': process_fnum(name.get('FNUM', ''), 'PASI'),
        'IORG': process_fnum(name.get('FNUM', ''), 'IORG'),
        'IGEN': process_fnum(name.get('FNUM', ''), 'IGEN'),
        'IRAD': process_fnum(name.get('FNUM', ''), 'IRAD'),
        'TI': process_text_field(name.get('TI', ''), max_length=5000),
        'AB': process_text_field(name.get('AB', ''), {'<br/>': '\n '}, remove_country_code=True, max_length=5000),
        'ADB': process_text_field(name.get('ADB', ''), {'<p>': ' ', '</p>': '\n '}, remove_country_code=True, max_length=5000),
        'TECD': process_text_field(name.get('TECD', ''), {'<br/>': ', '}, max_length=5000),
        'PERMALINK': name.get('PERMALINK', ''),
    }

# def clean_text(text):
#     """Удаляет первый номер (например, 'US9088787') из текста."""
#     return re.sub(r'^\([A-Z]{2,3}\d+\)<br/>', '', text)

def translate_text(text, target_language='ru'):
    """Переводит текст на указанный язык после удаления номера."""
    if text:
        #cleaned_text = clean_text(text)  # Удаляем номер перед переводом
        try:
            return GoogleTranslator(source='auto', target=target_language).translate(text)
        except Exception as e:
            print(f"Ошибка перевода: {e}")
            return text
    return text


def json_restore(json_file):
    """Основная функция для обработки JSON файла."""
    all_fields = {
        'true_EAPD': [], 'true_EPRD': [], 'true_PA': [], 'true_PASI': [], 'true_PTCC': [], 'true_LIC': [],
        'true_OPPI': [], 'true_STDN': [], 'true_IORG': [], 'true_IGEN': [], 'true_IRAD': [], 'true_PN_Count': [],
        'true_NPN': [], 'true_CT_COUNT': [], 'true_NPR': [], 'true_ICLM': [], 'true_CLMS_COUNT': [],
        'true_status': [], 'true_IC': [], 'true_TI': [], 'true_TI_ru': [], 'true_AB': [], 'true_AB_ru': [],
        'true_PER': [], 'true_ADB': [], 'true_ADB_ru': [], 'true_TECD': [], 'true_FAN': []#, 'true_PAVI': []
    }
    # test1 = []
    # test2 = []

    with open(json_file, encoding='utf-8') as file:
        data_json = json.load(file)
    #print(data_json)
    try:
        if data_json['nb'] == 0:
            print('Был получен пустой json')
            sys.exit(1)
    except Exception as e:
        if 'merged' in json_file:
            pass
        else:
            print(e)
            sys.exit(1)
    while len(data_json) != 0:
        if 'merged' in json_file:
            data = dict(data_json[0])
        else:
            data = data_json
            data_json = ['none']
        documents = data['documents']
        for name in documents:
            patent_info = extract_patent_restore(name)
            all_fields['true_FAN'].append(patent_info['FAN'])
            all_fields['true_EAPD'].append(patent_info['EAPD'])
            all_fields['true_EPRD'].append(patent_info['EPRD'])
            all_fields['true_PA'].append(patent_info['PA'])
            all_fields['true_PASI'].append(patent_info['PASI'])
            all_fields['true_PTCC'].append(patent_info['PTCC'])
            all_fields['true_LIC'].append(patent_info['LIC'])
            all_fields['true_OPPI'].append(patent_info['OPPI'])
            all_fields['true_STDN'].append(patent_info['STDN'])
            all_fields['true_IORG'].append(patent_info['IORG'])
            all_fields['true_IGEN'].append(patent_info['IGEN'])
            all_fields['true_IRAD'].append(patent_info['IRAD'])
           # all_fields['true_PAVI'].append(patent_info['PAVI'])
            all_fields['true_NPN'].append(patent_info['NPN'])
            all_fields['true_CT_COUNT'].append(len(patent_info['CTGN']) + len(patent_info['CTN']))
            all_fields['true_NPR'].append(patent_info['NPR'])
            try:
                all_fields['true_ICLM'].append(len(patent_info['ICLM'][0]))
                all_fields['true_CLMS_COUNT'].append(len(patent_info['CLMS'][0]) - len(patent_info['ICLM'][0]))
            except Exception as e:
                all_fields['true_ICLM'].append(len(patent_info['ICLM']))
                try:
                    all_fields['true_CLMS_COUNT'].append(len(patent_info['CLMS'][0]) - len(patent_info['ICLM']))
                except Exception:
                    all_fields['true_CLMS_COUNT'].append(len(patent_info['CLMS']) - len(patent_info['ICLM']))
            all_fields['true_IC'].append(patent_info['IC'])
            v_apl_info = patent_info['V_APL']
            if v_apl_info:
                all_fields['true_PN_Count'].append(len(v_apl_info.get('PN', [])))
                all_fields['true_status'].append(v_apl_info.get('STATUS', []))
            # test1.append('')
            # test2.append('')
            all_fields['true_TI'].append(patent_info['TI'])
            all_fields['true_TI_ru'].append(translate_text(patent_info['TI']))
            all_fields['true_AB'].append(patent_info['AB'])
            all_fields['true_AB_ru'].append(translate_text(patent_info['AB']))
            all_fields['true_ADB'].append(patent_info['ADB'])
            all_fields['true_ADB_ru'].append(translate_text(patent_info['ADB']))
            all_fields['true_PER'].append(patent_info['PERMALINK'])
            all_fields['true_TECD'].append(patent_info['TECD'])
            print('Another one')
            

        del data_json[0]

    # Используем len_num для ограничения длины всех полей
    len_num = len(all_fields['true_PA'])

    # Обрезаем все поля до длины len_num
    for key in all_fields:
        all_fields[key] = all_fields[key][:len_num]

    # Форматирование всех данных
    all_names = {
        'Questel unique family ID (FAN)': process_field_list(all_fields['true_FAN']),
        'Дата подачи патентной заявки': process_field_list(all_fields['true_EAPD']),
        'Дата приоритета': process_field_list(all_fields['true_EPRD']),
        'Компания (университет)': process_field_list(all_fields['true_PA']),
        # 'Теги основания «задачи летательных аппаратов»': process_field_list(test1),
        # 'Теги основания «методы и техники искусственного интеллекта»': process_field_list(test2),
        # 'Теги задач ОАК': process_field_list(test1),
        # 'Теги ИИ': process_field_list(test2),
        'Сила патентного семейства': process_field_list(all_fields['true_PASI']),
        'Страны (юрисдикции) патентования': process_field_list(all_fields['true_PTCC']),
        'Признаки лицензирования': process_field_list(all_fields['true_LIC']),
        'Признаки патентных споров': process_field_list(all_fields['true_OPPI']),
        'Признаки отнесения к патентам, существенным для стандарта': process_field_list(all_fields['true_STDN']),
        'Оригинальность патентного семейства': process_field_list(all_fields['true_IORG']),
        'Универсальность патентного семейства': process_field_list(all_fields['true_IGEN']),
        'Радикальность патентного семейства': process_field_list(all_fields['true_IRAD']),
        'Число патентных документов в патентном семействе': process_field_list(all_fields['true_PN_Count']),
        'Число патентов': process_field_list(all_fields['true_NPN']),
        'Число цитирований': process_field_list(all_fields['true_CT_COUNT']),
        'Число приоритетов': process_field_list(all_fields['true_NPR']),
        'Число независимых пунктов формулы': process_field_list(all_fields['true_ICLM']),
        'Число зависимых пунктов формулы': process_field_list(all_fields['true_CLMS_COUNT']),
        'Правовой статус': process_field_list(all_fields['true_status']),
        'Подклассы, группы и подгруппы Международной патентной классификации': process_field_list(all_fields['true_IC']),
        'Название на английском языке': process_field_list(all_fields['true_TI']),
        'Название на русском языке': process_field_list(all_fields['true_TI_ru']),
        'Реферат на английском языке': process_field_list(all_fields['true_AB']),
        'Реферат на русском языке': process_field_list(all_fields['true_AB_ru']),
        'Ссылки на оригинальный патентный документ': process_field_list(all_fields['true_PER']),
        'Описания преимуществ / недостатков предшествующего уровня техники на английском языке': process_field_list(all_fields['true_ADB']),
        'Описания преимуществ / недостатков предшествующего уровня техники на русском языке': process_field_list(all_fields['true_ADB_ru']),
        'Технологические концепты, автоматически выделяемые из текста патентных документов': process_field_list(all_fields['true_TECD'])
        #'Patent value': process_field_list(all_fields['true_PAVI'])
    }

    return all_names

company = 'Leonardo'
ticket, shard = login('Q_Ivan@rupto.ru', 'IiHeH8auR6UZ')
data_filepath = '/Users/igorkomissarov/ProjectOffice_FIPS Dropbox/Игорь Комиссаров/WorkPlace/bunch/' + company + '/' + company + '_FAN.xlsx'
fan_data = pd.read_excel(data_filepath)

    # Формирование строки запроса для всех FAN идентификаторов
fan_ids = ' OR '.join(fan_data['Questel unique family ID (FAN)'].astype(str))

    # Запрос данных через API и парсинг ответа
api_fan_company('FAN_company.json', fan_ids, 'aw', ticket, shard)

all = json_restore('/Users/igorkomissarov/ProjectOffice_FIPS Dropbox/Игорь Комиссаров/WorkPlace/bunch/FAN_company.json')
all = pd.DataFrame(all)
#all.to_excel('Restore.xlsx')
def make(all):
    #file_path = 'C:/website_orbit/website_rup/Restore.xlsx'
    df = all

    # Преобразуйте строки в числа с учетом запятой как десятичного разделителя
    for column in ['Оригинальность патентного семейства', 'Универсальность патентного семейства', 
                'Радикальность патентного семейства', 'Сила патентного семейства']:
        # Замените запятую на точку для преобразования в float
        df[column] = df[column].astype(str).str.replace(',', '.')
        # Преобразуйте в числовой формат
        df[column] = pd.to_numeric(df[column], errors='coerce')

    # Шаг 2: Сохраните DataFrame в новый Excel файл
    output_file_path = '/Users/igorkomissarov/ProjectOffice_FIPS Dropbox/Игорь Комиссаров/WorkPlace/bunch/' + company + '/Restore_' + company +'.xlsx'
    df.drop(columns=['Unnamed: 0'], inplace=True, errors='ignore')
    df.to_excel(output_file_path, index=False, engine='openpyxl')

    # Шаг 3: Используйте openpyxl для настройки локализованного формата ячеек
    wb = load_workbook(output_file_path)
    ws = wb.active
    # Настройте формат ячеек для нужных столбцов
    for column_letter in ['E', 'J', 'K', 'L']:  # Замените на нужные буквы столбцов
        for cell in ws[column_letter]:
            cell.number_format = '#,##0.00'  # Настройте формат с запятой

    # Сохраните изменения
    wb.save(output_file_path)
    print(f"Файл успешно сохранен с форматированием как {output_file_path}")

make(all)